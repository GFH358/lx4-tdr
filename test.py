

#-------------------------------------------------------------------
#xlwings开关文档 # 开发环境验证OK，可以关闭打开的excel文档窗口
import xlwings as xw
wb = xw.Book("任务目录汇总表.xlsx")                        # 连接到已打开的 Excel 文件
# 保存工作簿
wb.save()
wb.close()



#-------------------------------------------------------------------

# import psutil
#
# # 列出所有进程的PID
# for proc in psutil.process_iter():
#     try:
#         # 获取进程名称
#         process_name = proc.name()
#         # 获取进程ID
#         process_id = proc.pid
#         # 输出进程名称和进程ID
#         print(f"Process: {process_name} PID: {process_id}")
#     except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
#         pass
#-------------------------------------------------------------------
# # #10.23 获取excel打开状态
# import win32com.client
# import os

# def fileisopen1(filepath):
#     # 判断Excel文件是否已打开
#     # 如果目标工作簿已打开则返回TRUE，否则返回FALSE
#     excelapp = win32com.client.Dispatch("Excel.Application")
#     flag = False
#
#     num = excelapp.Workbooks.Count
#     print("num = %s" % num)
#     if num > 0:
#         for i in range(1,num+1):
#             temppath = os.path.abspath("..")
#             print(temppath)
#             realpth = excelapp.Workbooks(i).Path + "\\" + excelapp.Workbooks(i).Name
#             filepath = temppath + "\\" + filepath
#             print(filepath)
#             print(realpth)
#             if filepath == realpth:
#                 flag = True
#                 break
#     return flag
#
# path = "D:\\virtualE\\Lxproject\\004.任务目录生成器\\GitlabProject\\任务目录汇总表.xlsx"
# path = "任务目录汇总表.xlsx"
# flag = fileisopen1(path)
# if flag == 1:
#     print("open")
# else:
#     print("close")

# 可行，但每次都要循环整个电脑的运行的进程，会很慢
# import psutil
# import os
#
# def close_excel_file(file_path):
#     for proc in psutil.process_iter():
#         try:
#             for item in proc.open_files():
#                 if item.path == os.path.abspath(file_path):
#                     # proc.kill()
#                     return 1
#         except:
#             pass
#
# file_path = "..\\任务目录汇总表.xlsx"
# if close_excel_file(file_path):
#     print('{file_path} is open')
# else:
#     print('{file_path} is not open')
#-------------------------------------------------------------------
# # #10.22 获取路径测试
# import os
#
# current_path = os.getcwd()
#
# print("当前路径为：", current_path)


# #10.22 字符串拆分测试
# str ="123456789"
# if str[:2] == "12":
#     print("good")
# print("str:%s" % str[:2])



#-------------------------------------------------------------------
#10.19 路径选择窗口测试
# import tkinter as tk
# from  tkinter import filedialog
#
#
# window =tk.Tk()
# window.title('数据处理')
# window.geometry('600x380')
#
# # tk.Label(window ,text="文件路径：").place(x=50 ,y=50)
# var_name =tk.StringVar()  # 文件输入路径变量
#
# # tk.Label(window ,text="文件路径夹：").place(x=50 ,y=100)
# var_name2 =tk.StringVar()  # 文件夹输入路径变量
#
# entry_name =tk.Entry(window ,textvariable=var_name ,width=55)
# entry_name.place(x=120 ,y=50)
# entry_name2 =tk.Entry(window ,textvariable=var_name2 ,width=55)
# entry_name2.place(x=120 ,y=100)
#
#
# # 输入文件路径
# def selectPath_file():
#     path_ = filedialog.askdirectory()
#     var_name.set(path_)
#
#
# # 输入文件夹路径
# def selectPath_dir():
#     path_ = filedialog.askopenfilename(filetypes=[("数据表" ,[".xls", ".xlsx"])])
#     var_name2.set(path_)
#
# tk.Button(window, text = "路径选择1", command = selectPath_file).place(x=525 ,y=45)
# tk.Button(window, text = "路径选择2", command = selectPath_dir).place(x=525 ,y=95)
#
# #####画布#####
# try:
#     canvas =tk.Canvas(window ,width=100 ,height=120)
#     picture =tk.PhotoImage(file="angre.png")
#     image = canvas.create_image(0 ,0 ,anchor="nw" ,image=picture)
#     canvas.place(x=80 ,y=240)
# except:
#     pass
#
# def user_test():
#     # window.mainloop()
#     print("user_test")
#
# window.mainloop()
#-------------------------------------------------------------------
#python跨文件导入模块测试
# import sum_table
# list1 = [1,"STSSS",322,"杰理",147]
# list2 = [22,"STSSS",322,"杰理",258]
# sum_table.sumtable_add_data(list1, 'example.xlsx','客户问题')
# sum_table.sumtable_get_excel_sheet_list('example.xlsx')
#-------------------------------------------------------------------
#excel添加数据测试
# import xlwings as xw
# import os
# def check_excel_is_open():
#     # file_path = 'C:\\ProgramData\\xxx\\xxxx\\BmEquipmentLibrary'
#     file_name = 'example.xlsx'
#     # temp_file = file_path + '\\' + '~$' + file_name
#     temp_file = '~$' + file_name
#     if os.path.exists(temp_file):
#         print('excel已被打开')
#
# check_excel_is_open()
# list1 = [2,"STSSS",322,"杰理",5]
# # 连接到已打开的 Excel 文件
# wb = xw.Book("example.xlsx")
# sheet = wb.sheets["Sheet1"]#选择工作表
#
#
# nrows = sheet.used_range.last_cell.row  # 获取最大行数
# print(nrows)
# # 向单元格写入数据
# sheet.range("A1").value = "Hello, world!"
# str = 'A' + str(nrows+1)
# print(str)
# sheet.range(str).value = list1
# # 读取单元格数据
# print(sheet.range("A1").value)
# wb.save()
#保存原文件
#wb1.save(r'F:\xlwings\PresentData01.xlsx')
#另存为PresentData01.xlsx

#
# # 打开已存在的Excel文件
# workbook = openpyxl.load_workbook('example.xlsx')
#
# # 选择第一个工作表
# sheet = workbook.active(write_only=True)
#
# # 获取当前工作表的最后一行
# last_row = sheet.max_row
#
# # 新增一行的位置为最后一行的下一行
# new_row = last_row + 1
#
# # 在指定位置插入新行
# sheet.insert_rows(new_row)
#
# # 设置新行的数据
# sheet.cell(row=new_row, column=1).value = '中文'
# sheet.cell(row=new_row, column=2).value = 25
# sheet.cell(row=new_row, column=3).value = 'Male'
#
# # 保存修改后的Excel文件
# workbook.save('example.xlsx')

from openpyxl import Workbook

# # 打开Excel文件（以写入模式）
# wb = Workbook(write_only=True)
#
# # 获取活动工作表
# ws = wb.create_sheet()
#
# # 写入数据
# ws.append([1, 2, 3, 4])
#
# # 保存文件
# wb.save('output.xlsx')


#-------------------------------------------------------------------
# import tkinter
# from tkinter import ttk
#
# import sv_ttk
#
# root = tkinter.Tk()
#
# button = ttk.Button(root, text="Click me!")
# button.pack()
#
# # This is where the magic happens
# sv_ttk.set_theme("dark")
#
# root.mainloop()

# 透明窗口测试
# #Import the Tkinter Library
# from tkinter import *
#
# #Create an instance of Tkinter Frame
# win = Tk()
#
# #Set the geometry of window
# win.geometry("700x350")
#
# #Add a background color to the Main Window
# win.config(bg = '#add223')
#
# #Create a transparent window
# win.wm_attributes('-transparentcolor','#add223')
# win.mainloop()

#右下角三角标志
# from tkinter import *
# from tkinter.ttk import *
#
# root = Tk()
# root.geometry('150x100+888+444')
#
# si = Sizegrip()  # 创建一个Sizegrip组件
# si.pack(side=bottom, anchor=CENTER)  # 这种组件一般是位于窗体右下角
#
# print(si.keys())
#
# root.mainloop()

#-------------------------------------------------------------------
# 透明窗体测试
# import tkinter as tk
#
# root = tk.Tk()
#
# # 隐藏菜单栏和任务栏
# root.overrideredirect(1)
#
# # 设置窗口透明度为50%
# root.attributes("-alpha", 0.5)
#
# # 创建一个新的窗口作为标题栏
# title_bar = tk.Toplevel(root)
#
# # 将标题栏背景色设置为不透明的颜色
# title_bar.config(bg="#ECECEC")
#
# # 设置标题栏大小和位置
# title_bar.geometry(f"{root.winfo_screenwidth()}x30+0+0")
#
# # 运行主循环
# root.mainloop()