import xlwings as xw
import os
import openpyxl
from openpyxl.styles import Font
import win32com.client
import main

'''
TDR规定:
1. 所有TDR版本需要放在同一个文件夹[Lx4_TDR]下，命名格式为：Lx4_TDRv2.1
2. 总表文件需要放在[Lx4_TDR]文件夹内，与各版本文件夹同一级目录
注意事项：
该模块只能通过通过接口接收数据保证低耦合，不可import外部连接
'''
mode0_list = ["TDR版本", "序号", "开始时间", "芯片系列", "SDK版本", "客户", "问题描述", "现场", "生产"]
mode1_list = ["TDR版本", "序号", "开始时间", "芯片系列", "SDK版本", "同事", "问题描述", "现场", "生产"]
# 日常支持客户问题列表格式：[TDR版本，序号，开始时间，芯片系列，SDK版本，客户，问题描述，现场，生产]，一共9个数据
# ESO模式内部协助列表格式：[TDR版本，序号，开始时间，芯片系列，SDK版本，同事，问题描述，现场，生产]，一共9个数据

# TDR规定:
# 1. 所有TDR版本需要放在同一个文件夹[Lx4_TDR]下，命名格式为：Lx4_TDRv2.1
# 2. 总表文件需要放在[Lx4_TDR]文件夹内，与各版本文件夹同一级目录
tdr_version_unify_name = "Lx4_TDR"          #TDR统一命名,用于后续总表超链接
stable_path = '..\\'                # 总表路径 (相对路径，上一级目录)
stable_temp_path = '..\\~$'         # 总表临时文件路径，用于检查excel是否打开
test_excel_path = "..\\任务目录汇总表.xlsx"

# 共用常量，需要与main.py保持一致
task_directory_path = "TaskDirectory"
task_record_path = "TaskRecord"
stable_sheet_made0_name = "客户问题"        # 客户问题日常支持模式
stable_sheet_made1_name = "内部测试"        # 内部协助测试模式（ESO模式）
stable_wps_compatible_enable = 0    # wps兼容模式

# 检查总表excel文件是否存在，若不存在则创建
def stable_check_excel_exist(file_name):
    if os.path.exists(file_name):
        print("[stable]:excel <%s> is exist" % file_name)
    else:
        wb = openpyxl.Workbook()                    # 若excel不存在，此时必然不存在打开，使用<openpyxl>
        sheet = wb.active                           # 选择当前活动的工作表
        # 这里会有一个默认的sheet表，建议在这里添加使用说明。有空再写
        sheet.cell(row=1, column=1).value = "使用说明"

        mode0_sheet = wb.create_sheet(stable_sheet_made0_name)       # 创建工作表
        mode0_sheet.append(mode0_list)                              # 设置标题栏
        mode1_sheet = wb.create_sheet(stable_sheet_made1_name)       # 创建工作表
        mode1_sheet.append(mode1_list)                              # 设置标题栏
        wb.save(file_name)
        print('[warning] -> [stable]:no excel <%s>,created excel successfully' % file_name)

# 检查总表excel文件是否打开,用检查线程方法
'''
def stable_check_excel_open(file_name):
    # 判断Excel文件是否已打开
    # 如果目标工作簿已打开则返回1，否则返回0
    excelapp = win32com.client.Dispatch("Excel.Application")
    flag = 0

    num = excelapp.Workbooks.Count
    print("num = %s" % num)
    if num > 0:
        for i in range(1,num+1):
            temp_path = os.path.abspath("..")
            print(temp_path)
            realpth = excelapp.Workbooks(i).Path + "\\" + excelapp.Workbooks(i).Name
            file_path = temp_path + "\\" + file_name
            print(file_path)
            print(realpth)
            if file_path == realpth:
                flag = True
                break
    if flag == 1:
        print('[stable]:excel is open，use <xlwings>')
    else:
        print('[stable]:excel is close，use <openpyxl>')
    return flag

'''
# # 检查总表excel文件是否打开(WPS打开时也有临时文件，注意文件夹不要隐藏)
def stable_check_excel_open(file_name):
    temp_file = stable_temp_path + file_name    # 打开excel的时候会出现前缀是~$的临时文件
    if os.path.exists(temp_file):
        print('[stable]:excel is open，use <xlwings>')
        return 1
    else:
        print('[stable]:excel is close，use <openpyxl>')
        return 0


# 获取excel工作表列表 //不使用
'''
def stable_get_excel_sheet_list(file):
    if(stable_check_excel_open(file)):
        wb = xw.Book(file)                  # 连接到已打开的 Excel 文件
        sheet_list = wb.sheets
        print(type(sheet_list), sheet_list) # 返回的是class列表，有一些多余的东西
        num = len(wb.sheets)
        sheet_list = list(range(0, num))
        for i in range(0, num):
            temp_sheet = wb.sheets[i]
            sheet_list[i] = temp_sheet.name
        print(sheet_list)
        return sheet_list
    else:
        wb = openpyxl.load_workbook(file)   # 打开已存在的Excel文件
        sheet_list = wb.sheetnames
        print(type(sheet_list), sheet_list)
        return sheet_list
'''

# 获取超链接地址
def stable_get_hyperlink_path(mode,datalist,hyperlink_str,hyperlink_input_path):
    if hyperlink_input_path != "TDR_mode" and hyperlink_input_path != "dev_mode":
        if mode == 0:
            hyperlink_path = hyperlink_input_path + "\\" + task_directory_path + "\\" + hyperlink_str
        elif mode == 1:
            hyperlink_path = hyperlink_input_path + "\\" + task_record_path + "\\" + hyperlink_str + ".md"
        else:
            hyperlink_path = "error:01"
            print("[error] -> [stable]:mode error:01,return hyperlink path fail\n")
    else:
        version_str = tdr_version_unify_name + datalist[0]
        if hyperlink_input_path == "dev_mode":
            version_str = "lx4-tdr"  # 修改目录名，测试用
        if mode == 0:
            hyperlink_path = version_str + "\\" + task_directory_path + "\\" + hyperlink_str
        elif mode == 1:
            hyperlink_path = version_str + "\\" + task_record_path + "\\" + hyperlink_str + ".md"
        else:
            hyperlink_path = "error:02"
            print("[error] -> [stable]:mode error:02,return hyperlink path fail\n")
    return hyperlink_path


# excel文件打开条件下添加一行数据，使用 <xlwings>
def stable_add_data_in_open(datalist, file, sheet_name, hyperlink_str, hyperlink_path):
    wb = xw.Book(file)                        # 连接到已打开的 Excel 文件

    # 获取工作表是否存在 方法1：
    # sheet_list = wb.sheets
    # for list_unit_name in sheet_list:  # 检查工作表是否存在
    #     if sheet_name in str(list_unit_name):
    #         print(list_unit_name)
    #         flag = 1
    # if flag == 1:
    # 获取工作表列表
    num = len(wb.sheets)            # 获取工作表个数
    sheet_list = list(range(0, num))
    for i in range(0, num):         # 通过for循环把工作表名字一个个读出来写入列表
        temp_sheet = wb.sheets[i]
        sheet_list[i] = temp_sheet.name
    print(sheet_list)
    if sheet_name in sheet_list:
        print("[stable]:sheet <%s> is exist\n" % sheet_name)
    else:
        wb.sheets.add(sheet_name)             # 工作表不存在，新建工作表
        print("[warning] -> [stable]:no sheet <%s>,xlwings created sheet successfully \n" % sheet_name)
    sheet = wb.sheets[sheet_name]  # 选择工作表
    nrows = sheet.used_range.last_cell.row    # 获取最大行数
    strs = 'A' + str(nrows+1)                 # 设置起始单元格，最后一行+1表示新增数据行
    print("[stable]:write data list:%s" % datalist)
    sheet.range(strs).value = datalist        # 向单元格写入数据
    new_row = nrows + 1

    # 设置超链接
    if sheet_name == stable_sheet_made0_name:  # 日常支持模式
        task_directory_hyperlink_path = stable_get_hyperlink_path(0, datalist, hyperlink_str,hyperlink_path)  # 获取目录超链接地址
        task_record_hyperlink_path = stable_get_hyperlink_path(1, datalist, hyperlink_str,hyperlink_path)  # 获取记录超链接地址
        # serial_text = str(datalist[1])  # 要转成字符串，format里不能传0 # 已确认TDR传过来的数据都是str
        sheet.range(new_row, 7).value = '=HYPERLINK("{}","{}")'.format(task_directory_hyperlink_path, datalist[6])
        sheet.range(new_row, 2).value = '=HYPERLINK("{}","{}")'.format(task_record_hyperlink_path, datalist[1])
    elif sheet_name == stable_sheet_made1_name:  # 内部协助模式，没有md记录文件，只需要设置一个超链接
        task_directory_hyperlink_path = stable_get_hyperlink_path(0, datalist, hyperlink_str,hyperlink_path)  # 合成超链接地址
        sheet.range(new_row, 7).value = '=HYPERLINK("{}","{}")'.format(task_directory_hyperlink_path, datalist[6])
    else:
        print("[error] -> [stable]:mode error,set hyperlink fail\n")
    # print(sheet.range("A1").value)          # 读取单元格数据
    sheet.range(strs).expand('right').api.HorizontalAlignment = -4131 # 设置左对齐，统一格式
    wb.save()                                 # 保存修改后的Excel文件

# excel文件关闭条件下添加一行数据，使用 <openpyxl>
def stable_add_data_in_close(datalist,file,sheet_name,hyperlink_str,hyperlink_path):

    wb = openpyxl.load_workbook(file)         # 打开已存在的Excel文件
    sheet_list = wb.sheetnames                # 获取工作表列表
    if sheet_name in sheet_list:              # 检查工作表是否存在
        print("[stable]:sheet <%s> is exist\n" % sheet_name)
    else:
        wb.create_sheet(sheet_name)           # 工作表不存在，新建工作表
        print("[warning] -> [stable]:no sheet <%s>,openpyxl created sheet successfully \n" % sheet_name)
    sheet = wb[sheet_name]                    # 选择工作表
    last_row = sheet.max_row                  # 获取当前工作表的最后一行
    new_row = last_row + 1
    sheet.insert_rows(new_row)           # 在指定位置插入新行，最后一行+1表示新增数据行
    print("[stable]:write data list:%s" % datalist)
    sheet.append(datalist)                    # 设置新行的数据
    # 设置超链接
    if sheet_name == stable_sheet_made0_name:       # 日常支持模式
        task_directory_hyperlink_path = stable_get_hyperlink_path(0,datalist,hyperlink_str,hyperlink_path)  # 获取目录超链接地址
        task_record_hyperlink_path = stable_get_hyperlink_path(1, datalist,hyperlink_str,hyperlink_path)    # 获取记录超链接地址
        sheet.cell(row=new_row, column=7).value = '=HYPERLINK("{}","{}")'.format(task_directory_hyperlink_path, datalist[6])  # 以超链接方式写入表格单元
        sheet.cell(row=new_row, column=2).value = '=HYPERLINK("{}","{}")'.format(task_record_hyperlink_path, datalist[1])  # 以超链接方式写入表格单元
    elif sheet_name == stable_sheet_made1_name:     # 内部协助模式，没有md记录文件，只需要设置一个超链接
        task_directory_hyperlink_path = stable_get_hyperlink_path(0, datalist,hyperlink_str,hyperlink_path)  # 合成超链接地址
        sheet.cell(row=new_row, column=7).value = '=HYPERLINK("{}","{}")'.format(task_directory_hyperlink_path, datalist[6])  # 以超链接方式写入表格单元
    else:
        print("[error] -> [stable]:mode error,set hyperlink fail\n")
    # 设置超链接字体和下划线
    font = Font(
        name=None,  # 字体
        size=11,  # 字体大小
        color="0563C1",  # 字体颜色，用16进制rgb表示
        bold=False,  # 是否加粗，True/False
        italic=False,  # 是否斜体，True/False
        strike=None,  # 是否使用删除线，True/False
        underline='single',  # 下划线, 可选'singleAccounting', 'double', 'single', 'doubleAccounting'
    )
    if sheet_name == stable_sheet_made0_name:
        temp_cell = 'B' + str(new_row)
        sheet[temp_cell].font = font
    temp_cell = 'G' + str(new_row)
    sheet[temp_cell].font = font

    wb.save(file)                             # 保存修改后的Excel文件

# 数据预处理
def stable_data_pretreatment(datalist):
    output_list = datalist
    cnt = 0
    # 打印TDR发过来的数据包类型
    if 0:
        print("stable_data_pretreatment:datalist type:\n")
        for temp_data in datalist:
            print(type(temp_data))

    # 判断是否现场
    if output_list[7] == '0' or output_list[7] == 'N':
        output_list[7] = 'N'
    elif output_list[7] == '1' or output_list[7] == 'Y':
        output_list[7] = 'Y'
    else:
        output_list[7] = '*'
    # 判断是否生产
    if output_list[8] == '0' or output_list[8] == 'N':
        output_list[8] = 'N'
    elif output_list[8] == '1' or output_list[8] == 'Y':
        output_list[8] = 'Y'
    else:
        output_list[8] = '*'
    return output_list

# 关闭打开的excel文档
def stable_close_excel(file_path):
    wb = xw.Book(file_path)     # 连接到已打开的 Excel 文件
    wb.save()                   # 保存工作簿
    wb.close()                  #关闭excel文档

# WPS兼容模式
def stable_set_wps_compatible_enable(en):
    global stable_wps_compatible_enable
    stable_wps_compatible_enable = en
    print(f"[stable]:wps_compatible_set:{en}\n")

def stable_get_wps_compatible_enable():
    return stable_wps_compatible_enable


# excel添加一行新数据
def stable_add_data(datalist, file, sheet_name, hyperlink_str, hyperlink_path):
    print("[stable]:stable add data start\n")
    datalist = stable_data_pretreatment(datalist)       # 数据预处理
    file_path = stable_path + file

    stable_check_excel_exist(file_path)           # 检查总表是否存在，不存在则创建

    if(stable_get_wps_compatible_enable()):# WPS兼容模式
        if (stable_check_excel_open(file)):  # 若总表已打开，先关闭
            print("[stable]:wps_compatible_close_excel\n")
            stable_close_excel(file_path)
        stable_add_data_in_close(datalist, file_path, sheet_name, hyperlink_str, hyperlink_path)
    else:# 标准模式
        if(stable_check_excel_open(file)):  # 检查总表是否处于打开状态
            stable_add_data_in_open(datalist, file_path, sheet_name, hyperlink_str, hyperlink_path)
        else:
            # stable_add_data_in_open(datalist, file_path, sheet_name, hyperlink_str, hyperlink_path) # 兼容模式下，强行打开excel输入
            stable_add_data_in_close(datalist, file_path, sheet_name, hyperlink_str, hyperlink_path)

    print("[stable]:stable add data succeed \n")



# file = '任务目录汇总表3.xlsx'
# stable_add_data(mode0_list, file,'客户问题')
# stable_get_excel_sheet_list(test_excel_path)